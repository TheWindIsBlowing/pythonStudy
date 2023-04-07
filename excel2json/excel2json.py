# coding=utf-8
import collections
import os

import pandas as pd
import numpy as np
import json
import openpyxl


class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.int64):
            return int(obj)
        elif isinstance(obj, np.float64):
            return float(obj)
        else:
            return super(NumpyEncoder, self).default(obj)


def parse_cell(obj):
    if isinstance(obj, np.int64):
        return int(obj)
    elif isinstance(obj, np.float64):
        str_value = str(obj)
        str_list = str_value.split(".")
        if len(str_list) == 2 and int(str_list[1]) == 0:
            return int(str_list[0])
        return float(obj)
    elif isinstance(obj, object):
        return str(obj)
    else:
        print("obj is other type: ", obj)
        return obj


def excel2list(path):
    filter_keys = []
    json_keys = []
    data_table = pd.read_excel(path, engine="openpyxl")
    type_table = pd.read_excel(path, sheet_name="type", header=None, engine="openpyxl")

    data_table.info()

    for i in range(len(type_table.index.values)):
        filter_keys.append(type_table.iloc[i, 0])
        json_keys.append(type_table.iloc[i, 1])

    data_list = list()
    for i in data_table.index.values:
        data_dic = dict()
        idx = 0
        for j in data_table.columns.values:
            data_cell = data_table.loc[i, j]
            if j in filter_keys and not pd.isna(data_cell):
                the_key = json_keys[filter_keys.index(j)]
                data_dic[the_key] = parse_cell(data_cell)
            idx += 1
        data_list.append(data_dic)
    return data_list


def test_func1():
    data = pd.read_excel("./data/Steam_Buff.xlsx", sheet_name="type")
    # json_str = data.to_json(orient="records", force_ascii=False)
    # print(json_str)

    print(data.info())
    print(data.index.values)
    for i in range(len(data.index.values)):
        print(data.iloc[i, 1])

def check_float(numValue):
    strValue = str(numValue)
    strList = strValue.split('.')
    if len(strList) == 1:
        return numValue
    elif strList[1] == '0':
        return int(strList[0])
    else:
        return numValue


# string字符串如果以';'和','结尾的，把符号和后面的空格去掉
def process_str(str):
    result = str.rstrip(' ').rstrip(';').rstrip(',')
    result = result.replace('\\n', '\n')
    return result


def process_cell(cell_value, export_type):
    if export_type == '' or export_type != 'string' and export_type != 'bool':
        if isinstance(cell_value, str):
        #if type(cell_value).__name__ == "unicode" or type(cell_value).__name__ == "string":
            if cell_value.find('.') == -1:
                return int(cell_value)
            else:
                return float(cell_value)
        elif isinstance(cell_value, float):
            return check_float(cell_value)
        else:
            return cell_value
    elif export_type == 'bool':
        if cell_value == 0:
            return False
        else:
            return True
    else:
        if isinstance(cell_value, str):
            return process_str(cell_value)
        else:
            return process_str(str(check_float(cell_value)))


def test_func2():
    work_book = openpyxl.load_workbook("./data/Steam_Buff.xlsx", data_only=True)
    print(work_book)
    print(work_book.sheetnames)
    print(work_book["type"])
    data_table = work_book[work_book.sheetnames[0]]
    type_table = work_book["type"]

    for row in type_table.iter_rows():
        print(row[0].value)

    print("=======================")
    key_list = []
    alias_list = []
    type_list = []
    empty_data = True
    for row in type_table.iter_rows():
        if row[0].value:
            key_list.append(row[0].value)
            alias_list.append(row[1].value)
            if len(row) > 2:
                type_list.append(row[2].value.lower())
                if row[2].value.lower() != u'reject':
                    empty_data = False
            else:
                type_list.append(u'reject')
    print(key_list)
    print(alias_list)
    print(type_list)
    if not empty_data:
        index = []
        for i in range(len(key_list)):
            found = False
            for col in data_table.iter_cols():
                val = col[0].value
                if key_list[i] == val:
                    index.append(col[0].col_idx - 1)
                    found = True
                    break

            if not found:
                index.append(-1)

        data = []
        for row in data_table.iter_rows():
            if (row[0].value or row[0].value == 0):
                element = collections.OrderedDict()
                for j in range(len(key_list)):
                    row_v = row[index[j]]
                    if row_v.row == 1:
                        break
                    if row_v and (row_v.value or row_v.value == 0):
                        cell_value = row_v.value
                        if type_list[j] != u'reject' and cell_value != '' and index[j] != -1:
                            element[alias_list[j]] = process_cell(cell_value, type_list[j])
                    # r = i + 1
                    # c = index[j] + 1
                    # if type_list[j] != u'reject' and table.cell( row = r, column = c ).value != '' and index[j] != -1:
                    #     element[alias_list[j]] = process_cell(table.cell( row = r, column =  c ).value, type_list[j])
                if len(element.keys()) > 0:
                    data.append(element)

        data_json = json.dumps(data, sort_keys=False, indent=4, ensure_ascii=False)
        print('data_json: ' + data_json)


if __name__ == "__main__":
    excel_dir = "./data"
    result_dir = "./result"
    merge_json = dict()
    for dirpath, dirnames, filenames in os.walk(excel_dir):
        print("===========================================")
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            print(file_path)
            res = excel2list(file_path)
            # print(res)
            excel_name = os.path.basename(file_path).split(".")[0]
            merge_json[excel_name] = res
            json_name = os.path.join(result_dir, excel_name + ".json")
            print(json_name)
            json.dump(res, open(json_name, "w", encoding="utf-8"), ensure_ascii=False, indent=4)

        for dirname in dirnames:
            print("current dir: ", dirname)

    json.dump(merge_json, open("./result/Config.json", "w", encoding="utf-8"), ensure_ascii=False, indent=4)

    # result = excel2list("./data/Steam_Achieve.xlsx")
    # print(result)
    # data_json = json.dumps(result, cls=NumpyEncoder, ensure_ascii=False, indent=4)
    # print(data_json)
    # json.dump(result, open("./result/Steam_Achieve.json", "w", encoding="utf-8"), ensure_ascii=False, indent=4)

    # test_func2()

